#!/usr/bin/env python3
"""
为每个故事生成词汇Excel表格
词汇按故事中出现顺序排列，包含音标、词性、词义、例句
"""

import os
import re
import json
from pathlib import Path

# 配置
BASE_DIR = Path(__file__).parent.parent
RESULT_DIR = BASE_DIR / "result"
STATS_DIR = BASE_DIR / "statistics"
OUTPUT_DIR = BASE_DIR / "vocab_tables"  # 词汇表输出目录

def extract_vocab_from_html(html_file):
    """从HTML文件中提取词汇及例句（按出现顺序）"""
    with open(html_file, 'r', encoding='utf-8') as f:
        content = f.read()

    # 提取词汇标注：<span class="w">word(中文)📢</span>
    pattern = r'<span class="w">([a-zA-Z]+)\(([^)]+)\)📢</span>'
    matches = re.findall(pattern, content)

    # 去重但保持顺序
    vocab_list = []
    seen_words = set()

    for word, chinese in matches:
        if word not in seen_words:
            seen_words.add(word)
            vocab_list.append({
                'word': word,
                'chinese': chinese,
                'order': len(vocab_list) + 1
            })

    # 提取例句（为每个词汇找一个包含它的句子）
    for vocab in vocab_list:
        word = vocab['word']
        # 找到包含该词汇的段落
        paragraph_pattern = rf'<p>([^<]*<span class="w">{word}\([^)]+\)📢</span>[^<]*)</p>'
        paragraph_matches = re.findall(paragraph_pattern, content)

        if paragraph_matches:
            # 清理HTML标签
            sentence = paragraph_matches[0]
            sentence = re.sub(r'<[^>]+>', '', sentence)
            sentence = sentence.strip()
            # 限制长度
            if len(sentence) > 100:
                sentence = sentence[:100] + '...'
            vocab['sentence'] = sentence
        else:
            vocab['sentence'] = ''

    return vocab_list

def load_vocab_cache():
    """加载音标和词性缓存"""
    cache_file = BASE_DIR / "scripts" / "vocab_cache.json"
    if cache_file.exists():
        with open(cache_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def get_phonetic_and_pos(word, vocab_cache):
    """获取音标和词性"""
    if word in vocab_cache:
        return vocab_cache[word].get('phonetic', ''), vocab_cache[word].get('pos', '')

    # 尝试从API获取（使用缓存）
    import urllib.request
    import urllib.error

    url = f"https://api.dictionaryapi.dev/api/v2/entries/en/{word}"

    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode('utf-8'))

            if data and len(data) > 0:
                entry = data[0]
                phonetic = entry.get('phonetic', '')

                meanings = entry.get('meanings', [])
                pos_list = []
                for meaning in meanings[:2]:
                    pos = meaning.get('partOfSpeech', '')
                    if pos:
                        pos_list.append(pos)

                pos_cn = ', '.join(pos_list) if pos_list else ''

                # 中文词性转换
                pos_mapping = {
                    'noun': '名词',
                    'verb': '动词',
                    'adjective': '形容词',
                    'adverb': '副词',
                    'pronoun': '代词',
                    'preposition': '介词',
                    'conjunction': '连词',
                    'interjection': '感叹词',
                    'determiner': '限定词',
                    'exclamation': '感叹词'
                }

                pos_cn = ', '.join([pos_mapping.get(p, p) for p in pos_list])

                return phonetic, pos_cn
    except:
        pass

    return '', ''

def create_vocab_excel(story_num, story_name, vocab_list, vocab_cache):
    """为单个故事创建词汇Excel表格"""

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "词汇表"

        # 设置表头
        headers = ['序号', '单词', '音标', '词性', '中文释义', '例句']
        header_fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")
        header_font = Font(bold=True, size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 填充数据
        for i, vocab in enumerate(vocab_list, 1):
            word = vocab['word']
            chinese = vocab['chinese']
            sentence = vocab.get('sentence', '')

            # 获取音标和词性
            phonetic, pos = get_phonetic_and_pos(word, vocab_cache)

            row = i + 1
            ws.cell(row=row, column=1, value=vocab['order'])
            ws.cell(row=row, column=2, value=word)
            ws.cell(row=row, column=3, value=phonetic)
            ws.cell(row=row, column=4, value=pos)
            ws.cell(row=row, column=5, value=chinese)
            ws.cell(row=row, column=6, value=sentence)

        # 调整列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 60

        # 设置行高
        for row in range(1, len(vocab_list) + 2):
            ws.row_dimensions[row].height = 20

        # 保存文件
        output_file = OUTPUT_DIR / f"{story_num:02d}_{story_name}_词汇表.xlsx"
        wb.save(output_file)

        return True, output_file

    except Exception as e:
        return False, str(e)

def main():
    print("=" * 60)
    print("为每个故事生成词汇Excel表格")
    print("=" * 60)

    # 创建输出目录
    OUTPUT_DIR.mkdir(exist_ok=True)

    # 加载音标缓存
    print("\n加载音标缓存...")
    vocab_cache = load_vocab_cache()
    print(f"已加载 {len(vocab_cache)} 个词汇的音标和词性数据")

    # 获取所有学习版文件
    learn_files = sorted(RESULT_DIR.glob("*_学习版.html"))
    print(f"\n找到 {len(learn_files)} 个学习版文件")

    success_count = 0
    error_count = 0
    total_vocab = 0

    print("\n开始处理...")

    for i, html_file in enumerate(learn_files, 1):
        # 提取故事编号和名称
        filename = html_file.stem
        match = re.match(r'(\d+)_(.+)_学习版', filename)

        if match:
            story_num = int(match.group(1))
            story_name = match.group(2)
        else:
            story_num = i
            story_name = filename.replace('_学习版', '')

        print(f"\n{i}/100:3d} - Story {story_num:3d}: {story_name}")

        # 提取词汇
        vocab_list = extract_vocab_from_html(html_file)
        print(f"    提取词汇: {len(vocab_list)} 个")

        # 创建Excel
        success, result = create_vocab_excel(story_num, story_name, vocab_list, vocab_cache)

        if success:
            success_count += 1
            total_vocab += len(vocab_list)
            print(f"    生成文件: {result.name}")
        else:
            error_count += 1
            print(f"    错误: {result}")

    print("\n" + "=" * 60)
    print("生成完成！")
    print(f"成功: {success_count} 个文件")
    print(f"失败: {error_count} 个文件")
    print(f"总词汇数: {total_vocab} 个")
    print(f"输出目录: {OUTPUT_DIR}")
    print("=" * 60)

if __name__ == "__main__":
    main()