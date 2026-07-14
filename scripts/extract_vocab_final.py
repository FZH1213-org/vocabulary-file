#!/usr/bin/env python3
"""
词汇汇总脚本
从所有学习版HTML文件中提取词汇，统计出现次数，生成Excel表格
"""

import os
import re
import json
from collections import defaultdict
from pathlib import Path
import urllib.request
import urllib.error
import time

# 配置
BASE_DIR = Path(__file__).parent.parent
RESULT_DIR = BASE_DIR / "result"
STATS_DIR = BASE_DIR / "statistics"

def extract_vocab_from_html(html_file):
    """从HTML文件中提取词汇及例句"""
    with open(html_file, 'r', encoding='utf-8') as f:
        content = f.read()

    # 提取词汇标注：<span class="w">word(中文)📢</span>
    pattern = r'<span class="w">([a-zA-Z]+)\(([^)]+)\)📢</span>'
    matches = re.findall(pattern, content)

    # 提取包含该词汇的例句
    vocab_sentences = {}
    for word, chinese in matches:
        if word not in vocab_sentences:
            # 找到包含这个词的句子
            sentence_pattern = rf'[^<>]*<span class="w">{word}\({chinese}\)📢</span>[^<>]*'
            sentences = re.findall(sentence_pattern, content)
            if sentences:
                # 清理HTML标签
                sentence = sentences[0]
                sentence = re.sub(r'<[^>]+>', '', sentence)
                sentence = sentence.strip()
                vocab_sentences[word] = sentence

    return matches, vocab_sentences

def get_phonetic_and_pos(word):
    """从Free Dictionary API获取音标和词性"""
    url = f"https://api.dictionaryapi.dev/api/v2/entries/en/{word}"

    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode('utf-8'))

            if data and len(data) > 0:
                entry = data[0]
                phonetic = entry.get('phonetic', '')

                # 获取词性
                meanings = entry.get('meanings', [])
                pos_list = []
                for meaning in meanings[:2]:  # 只取前两个词性
                    pos = meaning.get('partOfSpeech', '')
                    if pos:
                        pos_list.append(pos)

                pos_cn = ', '.join(pos_list) if pos_list else ''

                # 中文词性转换
                pos_mapping = {
                    'noun': '名词',
                    'verb': '动词',
                    'adjective': '形容词',
                    'adverb': '副词',
                    'pronoun': '代词',
                    'preposition': '介词',
                    'conjunction': '连词',
                    'interjection': '感叹词',
                    'determiner': '限定词',
                    'exclamation': '感叹词'
                }

                pos_cn = ', '.join([pos_mapping.get(p, p) for p in pos_list])

                return phonetic, pos_cn
    except Exception as e:
        pass

    return '', ''

def main():
    print("=" * 60)
    print("词汇汇总脚本")
    print("=" * 60)

    # 获取所有学习版HTML文件
    html_files = sorted(RESULT_DIR.glob("*_学习版.html"))
    print(f"\n找到 {len(html_files)} 个学习版文件")

    # 提取所有词汇
    vocab_stats = defaultdict(lambda: {
        'count': 0,
        'chinese': '',
        'first_story': '',
        'sentence': ''
    })

    vocab_order = []  # 记录词汇首次出现的顺序
    story_num = 0

    print("\n正在提取词汇...")
    for html_file in html_files:
        story_num += 1
        story_name = html_file.stem.replace('_学习版', '')
        print(f"处理故事 {story_num}/100: {story_name}")

        matches, sentences = extract_vocab_from_html(html_file)

        for word, chinese in matches:
            if vocab_stats[word]['count'] == 0:
                # 首次出现
                vocab_order.append(word)
                vocab_stats[word]['first_story'] = story_name
                vocab_stats[word]['chinese'] = chinese
                vocab_stats[word]['sentence'] = sentences.get(word, '')

            vocab_stats[word]['count'] += 1

    print(f"\n共提取到 {len(vocab_stats)} 个不同词汇")
    print(f"总词汇标注数：{sum(v['count'] for v in vocab_stats.values())}")

    # 获取音标和词性（使用已有的汇总表作为缓存）
    print("\n正在获取音标和词性...")

    # 尝试从已有文件加载缓存
    cache_file = STATS_DIR / "vocab_cache.json"
    vocab_cache = {}
    if cache_file.exists():
        with open(cache_file, 'r', encoding='utf-8') as f:
            vocab_cache = json.load(f)
        print(f"从缓存加载了 {len(vocab_cache)} 个词汇的音标和词性")

    # 为每个词汇获取音标和词性
    vocab_data = []
    for i, word in enumerate(vocab_order, 1):
        if i % 10 == 0:
            print(f"处理进度: {i}/{len(vocab_order)}")

        stats = vocab_stats[word]

        # 从缓存获取或从API获取
        if word in vocab_cache:
            phonetic = vocab_cache[word].get('phonetic', '')
            pos = vocab_cache[word].get('pos', '')
        else:
            phonetic, pos = get_phonetic_and_pos(word)
            vocab_cache[word] = {'phonetic': phonetic, 'pos': pos}
            time.sleep(0.1)  # 避免请求过快

        vocab_data.append({
            '序号': i,
            '单词': word,
            '音标': phonetic,
            '词性': pos,
            '中文释义': stats['chinese'],
            '出现次数': stats['count'],
            '首次出现故事': stats['first_story'],
            '例句': stats['sentence']
        })

    # 保存缓存
    with open(cache_file, 'w', encoding='utf-8') as f:
        json.dump(vocab_cache, f, ensure_ascii=False, indent=2)

    # 生成Excel文件
    print("\n正在生成Excel文件...")

    try:
        import pandas as pd

        df = pd.DataFrame(vocab_data)

        # 保存为Excel
        excel_file = STATS_DIR / "词汇汇总表_扩充版.xlsx"
        df.to_excel(excel_file, index=False, engine='openpyxl')
        print(f"Excel文件已保存: {excel_file}")

        # 同时保存JSON格式
        json_file = STATS_DIR / "词汇汇总表_扩充版.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(vocab_data, f, ensure_ascii=False, indent=2)
        print(f"JSON文件已保存: {json_file}")

    except ImportError:
        print("pandas未安装，使用openpyxl直接生成Excel...")

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "词汇汇总"

        # 写入表头
        headers = ['序号', '单词', '音标', '词性', '中文释义', '出现次数', '首次出现故事', '例句']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 写入数据
        for row, data in enumerate(vocab_data, 2):
            ws.cell(row=row, column=1, value=data['序号'])
            ws.cell(row=row, column=2, value=data['单词'])
            ws.cell(row=row, column=3, value=data['音标'])
            ws.cell(row=row, column=4, value=data['词性'])
            ws.cell(row=row, column=5, value=data['中文释义'])
            ws.cell(row=row, column=6, value=data['出现次数'])
            ws.cell(row=row, column=7, value=data['首次出现故事'])
            ws.cell(row=row, column=8, value=data['例句'])

        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 50

        # 保存
        excel_file = STATS_DIR / "词汇汇总表_扩充版.xlsx"
        wb.save(excel_file)
        print(f"Excel文件已保存: {excel_file}")

    print("\n" + "=" * 60)
    print("汇总完成！")
    print(f"总词汇数: {len(vocab_stats)}")
    print(f"总标注数: {sum(v['count'] for v in vocab_stats.values())}")
    print("=" * 60)

if __name__ == "__main__":
    main()